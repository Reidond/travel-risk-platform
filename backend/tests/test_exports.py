"""Export smoke tests: xlsx parses back, PDF renders with embedded fonts."""

import io

import pytest
from openpyxl import load_workbook

from tests.conftest import CRITERIA_CODES, create_control_region

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def run(client) -> dict:
    region_id = create_control_region(client)
    return client.post("/api/evaluations", json={"region_ids": [region_id]}).json()


@pytest.mark.parametrize("lang", ["uk", "en"])
def test_xlsx_export(client, run, lang):
    response = client.get(
        f"/api/evaluations/{run['id']}/export", params={"format": "xlsx", "lang": lang}
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX_MIME)
    assert "attachment" in response.headers["content-disposition"]
    assert len(response.content) > 0

    workbook = load_workbook(io.BytesIO(response.content))
    assert len(workbook.sheetnames) == 2  # summary + one region sheet
    summary = workbook[workbook.sheetnames[0]]
    rows = list(summary.iter_rows(values_only=True))
    assert len(rows) == 2  # header + one region
    expected_region = "Контрольний регіон" if lang == "uk" else "Control region"
    assert rows[1][0] == expected_region
    assert rows[1][1] == 3  # n
    assert rows[1][2] == 0.85  # xi
    assert rows[1][3] == "D5"
    assert rows[1][4] == 70.0  # delta = mean chi
    assert str(rows[1][9]).startswith("R5")

    region_sheet = workbook[workbook.sheetnames[1]]
    individual_rows = list(region_sheet.iter_rows(values_only=True))
    assert len(individual_rows) == 4  # header + 3 respondents
    assert individual_rows[1][1] == "e1"  # ext_id column
    header = individual_rows[0]
    assert any("G1" in str(cell) for cell in header)


@pytest.mark.parametrize("lang", ["uk", "en"])
def test_pdf_export(client, run, lang):
    response = client.get(
        f"/api/evaluations/{run['id']}/export", params={"format": "pdf", "lang": lang}
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    content = response.content
    assert content.startswith(b"%PDF")
    # Embedded TrueType font program (Cyrillic-capable DejaVu).
    assert b"FontFile2" in content
    assert b"DejaVu" in content


def test_xlsx_export_neutralizes_formula_injection(client):
    """User-controlled strings starting like formulas must export as inert text."""
    payload = '=HYPERLINK("http://evil")'
    region = client.post(
        "/api/regions",
        json={"name_uk": "=2+5", "name_en": "=2+5", "xi": 0.5, "delta_level": "D3"},
    ).json()
    response = client.post(
        f"/api/regions/{region['id']}/respondents",
        json={"ext_id": payload, "ratings": dict.fromkeys(CRITERIA_CODES, 3)},
    )
    assert response.status_code == 201, response.text
    run = client.post("/api/evaluations", json={"region_ids": [region["id"]]}).json()

    export = client.get(f"/api/evaluations/{run['id']}/export", params={"format": "xlsx"})
    workbook = load_workbook(io.BytesIO(export.content))
    summary_cell = workbook[workbook.sheetnames[0]]["A2"]
    region_cell = workbook[workbook.sheetnames[1]]["B2"]
    for cell, original in ((summary_cell, "=2+5"), (region_cell, payload)):
        assert cell.data_type != "f"  # never a live formula
        assert cell.value == "'" + original  # OWASP quote-prefix neutralization


def test_export_validation(client, run):
    assert (
        client.get(f"/api/evaluations/{run['id']}/export", params={"format": "docx"}).status_code
        == 422
    )
    assert client.get("/api/evaluations/9999/export").status_code == 404
