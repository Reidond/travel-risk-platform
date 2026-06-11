"""Questionnaire import: .xlsx / .csv in the «Експертне оцінювання.xlsx» format.

Columns (positional): respondent id, «Рік відпустки (перебування): »,
«Місяць відпустки (перебування): », «Відвідуваний район: » (oblast in
parentheses), «Тип розміщення: », K1..K17 mapped positionally to the active
criteria codes. Any extra columns (U, м) are ignored. Bad rows are collected
as row-level errors and never abort the whole file.
"""

import csv
import io
import re
import zipfile
from dataclasses import dataclass, field
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import models
from app.core_bridge import criteria_codes, load_groups
from app.paths import DEMO_XLSX
from app.translit import transliterate

_OBLAST_RE = re.compile(r"\(([^()]+)\)")

#: Hard cap on data rows per uploaded file (DoS guard for the public demo).
MAX_IMPORT_ROWS = 10_000

#: Demo regions with the article's DM inputs (contract: /api/import/demo).
DEMO_REGIONS: list[dict[str, Any]] = [
    {
        "name_uk": "Закарпатська область",
        "name_en": "Zakarpattia Oblast",
        "xi": 0.85,
        "delta_level": "D5",
    },
    {
        "name_uk": "Івано-Франківська область",
        "name_en": "Ivano-Frankivsk Oblast",
        "xi": 0.78,
        "delta_level": "D4",
    },
    {"name_uk": "Львівська область", "name_en": "Lviv Oblast", "xi": 0.88, "delta_level": "D4"},
]


@dataclass
class ParsedRow:
    row_number: int
    ext_id: str | None
    year: str | None
    month: str | None
    rayon: str | None
    accommodation: str | None
    rating_cells: list[Any]


@dataclass
class ImportOutcome:
    imported: int = 0
    skipped: int = 0
    created_regions: list[models.Region] = field(default_factory=list)
    errors: list[dict[str, Any]] = field(default_factory=list)


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _check_row_cap(rows: list[ParsedRow], max_rows: int | None) -> None:
    if max_rows is not None and len(rows) > max_rows:
        raise HTTPException(
            status_code=422,
            detail=f"file has more than {max_rows} data rows; split it into smaller uploads",
        )


def parse_xlsx(data: bytes, n_ratings: int, max_rows: int | None = None) -> list[ParsedRow]:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (zipfile.BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise HTTPException(
            status_code=422,
            detail="file is not a valid .xlsx workbook (corrupt or wrong format)",
        ) from exc
    worksheet = workbook.active
    if worksheet is None:
        return []
    rows: list[ParsedRow] = []
    for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_number == 1:  # header row
            continue
        rows.append(_parse_values(row_number, list(values), n_ratings))
        _check_row_cap(rows, max_rows)
    workbook.close()
    return rows


def parse_csv(data: bytes, n_ratings: int, max_rows: int | None = None) -> list[ParsedRow]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(
            status_code=422,
            detail="file is not valid UTF-8 text; re-export the CSV as UTF-8",
        ) from exc
    rows: list[ParsedRow] = []
    for row_number, values in enumerate(csv.reader(io.StringIO(text)), start=1):
        if row_number == 1:  # header row
            continue
        rows.append(_parse_values(row_number, list(values), n_ratings))
        _check_row_cap(rows, max_rows)
    return rows


def _parse_values(row_number: int, values: list[Any], n_ratings: int) -> ParsedRow:
    padded = values + [None] * max(0, 5 + n_ratings - len(values))
    return ParsedRow(
        row_number=row_number,
        ext_id=_cell_str(padded[0]),
        year=_cell_str(padded[1]),
        month=_cell_str(padded[2]),
        rayon=_cell_str(padded[3]),
        accommodation=_cell_str(padded[4]),
        rating_cells=padded[5 : 5 + n_ratings],
    )


def _is_empty(row: ParsedRow) -> bool:
    return (
        row.ext_id is None
        and row.rayon is None
        and all(_cell_str(cell) is None for cell in row.rating_cells)
    )


def extract_oblast(rayon: str | None) -> str | None:
    if not rayon:
        return None
    match = _OBLAST_RE.search(rayon)
    if match is None:
        return None
    oblast = match.group(1).strip()
    return oblast or None


def _parse_ratings(row: ParsedRow, codes: list[str]) -> dict[str, int]:
    """Validate rating cells; raises ValueError with a row-level message."""
    ratings: dict[str, int] = {}
    for code, cell in zip(codes, row.rating_cells, strict=True):
        if cell is None or (isinstance(cell, str) and not cell.strip()):
            raise ValueError(f"missing rating for criterion {code}")
        try:
            value = int(str(cell).strip())
        except ValueError:
            raise ValueError(f"invalid rating {cell!r} for criterion {code}") from None
        if not 1 <= value <= 5:
            raise ValueError(f"rating {value} for criterion {code} is out of range 1-5")
        ratings[code] = value
    return ratings


def parse_upload(filename: str, data: bytes, n_ratings: int) -> list[ParsedRow]:
    lowered = filename.lower()
    if lowered.endswith(".xlsx"):
        return parse_xlsx(data, n_ratings, max_rows=MAX_IMPORT_ROWS)
    if lowered.endswith(".csv"):
        return parse_csv(data, n_ratings, max_rows=MAX_IMPORT_ROWS)
    raise HTTPException(status_code=422, detail="unsupported file type; expected .xlsx or .csv")


def import_rows(
    session: Session,
    rows: list[ParsedRow],
    target_region: models.Region | None,
    region_by_oblast: dict[str, models.Region] | None = None,
    auto_create: bool = True,
) -> ImportOutcome:
    """Insert respondents row by row, collecting per-row errors."""
    groups = load_groups(session)
    codes = criteria_codes(groups)
    outcome = ImportOutcome()
    oblast_map: dict[str, models.Region] = dict(region_by_oblast or {})

    for row in rows:
        if _is_empty(row):
            continue
        try:
            ratings = _parse_ratings(row, codes)
        except ValueError as exc:
            outcome.skipped += 1
            outcome.errors.append({"row": row.row_number, "message": str(exc)})
            continue

        region = target_region
        if region is None:
            oblast = extract_oblast(row.rayon)
            if oblast is None:
                outcome.skipped += 1
                outcome.errors.append(
                    {
                        "row": row.row_number,
                        "message": "cannot determine oblast from the rayon column",
                    }
                )
                continue
            region = oblast_map.get(oblast)
            if region is None:
                region = session.scalars(
                    select(models.Region).where(models.Region.name_uk == oblast)
                ).first()
            if region is None:
                if not auto_create:
                    outcome.skipped += 1
                    outcome.errors.append(
                        {"row": row.row_number, "message": f"unknown region {oblast!r}"}
                    )
                    continue
                region = models.Region(name_uk=oblast, name_en=transliterate(oblast))
                session.add(region)
                session.flush()
                outcome.created_regions.append(region)
            oblast_map[oblast] = region

        session.add(
            models.Respondent(
                region_id=region.id,
                ext_id=row.ext_id,
                year=row.year,
                month=row.month,
                accommodation=row.accommodation,
                ratings=ratings,
            )
        )
        outcome.imported += 1

    return outcome


def import_demo(session: Session) -> tuple[ImportOutcome, list[models.Region]]:
    """Seed the bundled 327-respondent demo dataset with the article's DM inputs.

    Idempotent: raises 409 if any demo region already has respondents.
    """
    existing: dict[str, models.Region | None] = {}
    occupied: list[str] = []
    for spec in DEMO_REGIONS:
        name_uk = str(spec["name_uk"])
        region = session.scalars(
            select(models.Region).where(models.Region.name_uk == name_uk)
        ).first()
        if region is not None and len(region.respondents) > 0:
            occupied.append(name_uk)
        existing[name_uk] = region
    if occupied:
        raise HTTPException(
            status_code=409,
            detail=(
                "demo dataset already imported; regions with respondents: " + ", ".join(occupied)
            ),
        )

    regions: dict[str, models.Region] = {}
    created: list[models.Region] = []
    for spec in DEMO_REGIONS:
        name_uk = str(spec["name_uk"])
        region = existing[name_uk]
        if region is None:
            region = models.Region(
                name_uk=name_uk,
                name_en=str(spec["name_en"]),
                xi=spec["xi"],
                delta_level=spec["delta_level"],
            )
            session.add(region)
            session.flush()
            created.append(region)
        else:
            region.xi = spec["xi"]  # type: ignore[assignment]
            region.delta_level = spec["delta_level"]  # type: ignore[assignment]
        regions[name_uk] = region

    groups = load_groups(session)
    n_ratings = len(criteria_codes(groups))
    rows = parse_xlsx(DEMO_XLSX.read_bytes(), n_ratings)
    outcome = import_rows(
        session,
        rows,
        target_region=None,
        region_by_oblast=regions,
        auto_create=False,
    )
    outcome.created_regions = created
    return outcome, created
