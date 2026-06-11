"""Full-dataset verification of the tourism-risk core (plan stage 2).

Run from the repository root:

    uv run --project core python core/scripts/verify_dataset.py

Parses the 327-respondent survey file ``docs/Експертне оцінювання.xlsx``
(columns: id, year, month, rayon with oblast in parentheses, accommodation
type, K1..K17, U, м), groups respondents by oblast (Закарпатська -> R1,
Івано-Франківська -> R2, Львівська -> R3), feeds the K values directly as
tau (the file uses positive coding, MATH_SPEC §5 — no inversion), and runs
``evaluate_region`` per region with the article's DM inputs:

    Xi = 0.85 / 0.78 / 0.88 and Delta = D5 / D4 / D4 for R1 / R2 / R3,

under ``DEFAULT_CONFIG``. Prints per region: n, the r*(e) distribution,
delta, phi, m_S, omega, mu_R and the risk class, plus row-validation
anomalies (ratings outside 1..5 or missing values), and emits a JSON block
for downstream tooling.
"""

from __future__ import annotations

import json
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from tourism_risk import (
    DEFAULT_CONFIG,
    ExpertSafetyLevel,
    RegionResult,
    RiskTerm,
    evaluate_region,
)

DATASET_PATH = Path(__file__).resolve().parents[2] / "docs" / "Експертне оцінювання.xlsx"

#: Number of K columns (criteria) expected per row.
CRITERIA_COUNT = 17

#: First K column (0-based) in the worksheet row tuple.
FIRST_K_COLUMN = 5


@dataclass(frozen=True, slots=True)
class RegionSpec:
    """Article DM inputs for one demonstration region."""

    code: str
    oblast: str
    name_en: str
    xi: float
    delta_level: ExpertSafetyLevel
    expected_n: int


REGIONS: tuple[RegionSpec, ...] = (
    RegionSpec("R1", "Закарпатська", "Zakarpattia", 0.85, ExpertSafetyLevel.D5, 209),
    RegionSpec("R2", "Івано-Франківська", "Ivano-Frankivsk", 0.78, ExpertSafetyLevel.D4, 41),
    RegionSpec("R3", "Львівська", "Lviv", 0.88, ExpertSafetyLevel.D4, 77),
)


@dataclass(frozen=True, slots=True)
class Row:
    """One validated respondent row from the survey file."""

    respondent_id: str
    oblast: str
    ratings: tuple[int, ...]


def extract_oblast(rayon_cell: object) -> str:
    """Pull the oblast name out of 'Rayon (Oblast область)'."""
    text = str(rayon_cell or "")
    if "(" not in text or ")" not in text:
        raise ValueError(f"no oblast in parentheses: {text!r}")
    inner = text[text.index("(") + 1 : text.rindex(")")]
    return inner.replace("область", "").strip()


def parse_rows(path: Path) -> tuple[list[Row], list[str]]:
    """Parse and validate all respondent rows; return (rows, anomalies)."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[Row] = []
    anomalies: list[str] = []
    for index, values in enumerate(sheet.iter_rows(values_only=True)):
        if index == 0:  # header
            continue
        respondent_id = str(values[0] or f"row{index}")
        try:
            oblast = extract_oblast(values[3])
        except ValueError as error:
            anomalies.append(f"{respondent_id}: {error}")
            continue
        raw = values[FIRST_K_COLUMN : FIRST_K_COLUMN + CRITERIA_COUNT]
        row_errors = [
            f"K{position + 1}={cell!r}"
            for position, cell in enumerate(raw)
            if not (isinstance(cell, int) and not isinstance(cell, bool) and 1 <= cell <= 5)
        ]
        if len(raw) != CRITERIA_COUNT or any(cell is None for cell in raw):
            row_errors.append("missing rating cells")
        if row_errors:
            anomalies.append(f"{respondent_id}: invalid ratings ({', '.join(row_errors)})")
            continue
        rows.append(Row(respondent_id, oblast, tuple(int(cell) for cell in raw)))  # type: ignore[arg-type]
    workbook.close()
    return rows, anomalies


def term_distribution(result: RegionResult) -> dict[str, int]:
    """r*(e) counts over L, BA, A, AA, H (all terms, including zero counts)."""
    counts = Counter(respondent.risk_term for respondent in result.respondents)
    return {term.name: counts.get(term, 0) for term in RiskTerm}


def main() -> int:
    if not DATASET_PATH.exists():
        print(f"dataset not found: {DATASET_PATH}", file=sys.stderr)
        return 1

    rows, anomalies = parse_rows(DATASET_PATH)
    print(f"Dataset: {DATASET_PATH}")
    print(f"Parsed rows: {len(rows)} valid, {len(anomalies)} anomalies")
    for anomaly in anomalies:
        print(f"  ANOMALY: {anomaly}")

    by_oblast: dict[str, list[Row]] = {}
    for row in rows:
        by_oblast.setdefault(row.oblast, []).append(row)
    unexpected = set(by_oblast) - {spec.oblast for spec in REGIONS}
    if unexpected:
        print(f"  ANOMALY: unexpected oblasts: {sorted(unexpected)}")

    summary: dict[str, dict[str, object]] = {}
    for spec in REGIONS:
        regional_rows = by_oblast.get(spec.oblast, [])
        flag = "" if len(regional_rows) == spec.expected_n else f"  (EXPECTED {spec.expected_n}!)"
        result = evaluate_region(
            respondent_ratings=[row.ratings for row in regional_rows],
            xi=spec.xi,
            delta_level=spec.delta_level,
            config=DEFAULT_CONFIG,
        )
        distribution = term_distribution(result)
        print(f"\n=== {spec.code} {spec.name_en} ({spec.oblast} область) ===")
        print(f"  n            = {len(regional_rows)}{flag}")
        print(f"  Xi (DM)      = {spec.xi},  Delta (DM) = {spec.delta_level.name}")
        print(f"  r* counts    = {distribution}")
        print(f"  delta        = {result.delta:.4f}")
        print(f"  phi          = {result.phi:.6f}")
        print(f"  m_S          = {result.m_s:.6f}")
        print(f"  omega        = {result.omega:.4f}")
        print(f"  mu_R         = {result.mu:.6f}")
        print(f"  risk class   = {result.risk_class.name}")
        summary[spec.code] = {
            "oblast": spec.oblast,
            "n": len(regional_rows),
            "xi": spec.xi,
            "delta_level": spec.delta_level.name,
            "r_star": distribution,
            "delta": round(result.delta, 4),
            "phi": round(result.phi, 6),
            "m_s": round(result.m_s, 6),
            "omega": round(result.omega, 4),
            "mu": round(result.mu, 6),
            "risk_class": result.risk_class.name,
        }

    print("\nJSON summary:")
    print(json.dumps({"anomalies": anomalies, "regions": summary}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
